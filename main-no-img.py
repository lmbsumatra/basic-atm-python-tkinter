from tkinter import *
import tkinter.messagebox
import openpyxl
from openpyxl import Workbook
from os import path

################################################################
# ATM Program - Group 9                                        #
# Author/s: Sumatra, Verdaluza, Regala, Vertudazo, Dimaano     #
# Description: An ATM program with registration, balance       #
#              inquiry withdraw,deposit, change password, and  #
#              log in functions. The code runs with GUI        #
# Date Accomplished: Jan 24, 2023; 8:18pm                      #
################################################################


root = Tk()

class ATMapp:

    # check if file exists, create one if not
    if (path.exists('g9db.xlsx')):
        xlfile = "g9db.xlsx"    # file name
    else:
        xl = Workbook()
        xl.save('g9db.xlsx')
        xlfile = "g9db.xlsx" 

    green = '#3aa15c' # color
    red = '#d63a3a'   # color
    white = 'white'
    gray = '#a5a5a5'
    font = "Times New Roman"    # font
    

    # where user's info will be appended, can be use in whole class! ;O
    user_infos = []

#################### login page, first page ####################
    def __init__(self, master):
        # initializing frame
        frame = Frame(root,width = 600,height = 500)
        frame.place(x=0,y=0)
        master.geometry('500x400')
        master.resizable(0,0)
        master.title("DSA Bank")
        
        # page login label
        label = Label(master, width = 43, text = "Log in", pady = 5, font = (self.font, 16, 'bold'), background = self.red, foreground = "white")
        label.place(x = 10, y = 45)

        # username label
        unlbl = Label(master, text = "Account ID:", font = (self.font, 13))
        unlbl.place(x = 120, y = 120)

        # entry box for username
        input_username = Entry(width = 14,font = ('Times New Roman', 20), bd = 0 )
        input_username.place(x = 235,y = 115)

        # password label
        pwlbl = Label(master, text = "Password:", font = (self.font, 14))
        pwlbl.place(x = 120, y = 175)

        # entry box for password
        input_userpass = Entry(width = 14,show = "*",font = ('Times New Roman', 20), bd = 0 )
        input_userpass.place(x = 235,y = 170)

        # log in button
        login_btn = Button(master, text = "Log in", width = 5, font = ('Times New Roman', 15), bg = self.green, fg = self.white,
        borderwidth = 0, command = lambda: self.importdata(input_username, input_userpass))
        login_btn.place(x = 345, y = 220)

        # register button
        register_btn = Button(master, text = "Register", width = 8, font = ('Times New Roman', 15), bg = self.gray, fg = self.white,
        borderwidth = 0, command = lambda: self.registerPg(master))
        register_btn.place(x = 317, y = 265)


    def getdata(self, user, row):
        # user's specific row in database
        global row_no
        row_no = row
            # user from database append to user_infos to use in whole program
        for info in user:
            self.user_infos.append(info)

        # successful log in! go to option page :)
        self.optionPg(root)

    def importdata(self, get_username, get_password):   
        # counter for row
        counter = 0

        # where row of data will be appended
        db_user = [] 

        # from login page user inputs: username & password
        # using .get() function to turn into value/s
        input_un = get_username.get()
        input_pw = get_password.get()

        # opening a excel py
        xl = openpyxl.load_workbook(self.xlfile)
        data = xl.active
        rws = data.iter_rows(min_row = 1, max_col= 7, max_row = None, min_col = 1, values_only = True)
        
        # iterating rows in excel
        for i, j, k, l, m, n, o in rws:
            counter = counter + 1
            if i == input_un:
                # if user is true, append data to db_user list
                db_user.extend([i, j, k, l, m, n, o])
                break
        # checking if login input pw == database pw
        if (len(db_user)) == 0 or (len(input_pw)) == 0:
            tkinter.messagebox.showinfo('Failed','Invalid Login Username or Password')
        elif (len(db_user[2])) > 0:
            userpass = str(db_user[2])
            if input_pw == userpass:
                # if == call getdata() to save THAT row/user info's
                self.getdata(db_user, counter)
            else:
                tkinter.messagebox.showinfo('Failed','Invalid Login Username or Password')
        else:
            tkinter.messagebox.showinfo('Failed','Invalid Login Username or Password')
        row_no = counter
        root.mainloop()

#################### registration page ####################
    def registerPg(self, master):
        # initializing frame
        frame = Frame(root,width=600,height=500)
        frame.place(x=0,y=0)
        master.geometry('500x400')
        master.resizable(0,0)

        # page login label
        myLabel = Label(master, width = 43, text = "Registration Form", pady = 5, font = (self.font, 16, 'bold'), background = self.red, foreground = "white")
        myLabel.place(x = 10, y = 45)

        #Id number
        LblId = Label(master, text="Account ID",width=20,font=(self.font, 14))  
        LblId.place(x=63,y=95) 

        Id_entry = IntVar        
        Id_entry = Entry(master)  
        Id_entry.place(x=250,y=95,width=150)   
   
        #fullname        
        LblFullname = Label(master, text="Full Name",width=20,font=(self.font, 14))  
        LblFullname.place(x=60,y=125) 

        Fullname_entry = StringVar        
        Fullname_entry = Entry(master)  
        Fullname_entry.place(x=250,y=125,width=150)  
        
        #password        
        LblPassword = Label(master, text="Password",width=20,font=(self.font, 14))  
        LblPassword.place(x=56,y=155) 
        
        Password_entry = Entry(master, show="*")  
        Password_entry.place(x=250,y=155,width=150)  

        #contact number
        LblNum = Label(master, text="Contact number",width=20,font=(self.font, 14))  
        LblNum.place(x=80,y=185) 

        Num_entry = IntVar         
        Num_entry = Entry(master)  
        Num_entry.place(x=250,y=185,width=150) 
        
        # age        
        LblAge = Label(master, text="Age",width=20,font=(self.font, 14))  
        LblAge.place(x=43,y=210) 
                
        Age_entry = Entry(master)  
        Age_entry.place(x=250,y=215,width=150) 

        #balance        
        LblBal = Label(master, text="Balance",width=20,font=(self.font, 14))  
        LblBal.place(x=50,y=245) 

        Bal_entry = DoubleVar()
        Bal_entry = Entry(master)  
        Bal_entry.place(x=250,y=245,width=150) 

        
        #gender        
        LblGender = Label(master, text="Gender",width=20,font=(self.font, 14))  
        LblGender.place(x=47,y=275)
        
        Gender = StringVar()
        Radiobutton(master, text="Male",padx = 10, variable=Gender, value='male').place(x=243,y=275)
        Radiobutton(master, text="Female",padx = 20, variable=Gender, value='female').place(x=310,y=275)  

        def PopUp():
            if toggle.get() == 1:
                 tkinter.messagebox.showinfo("Terms & Condition","Ticking the box is allowing us to collect your personal infos.")
    
        #terms and condition
        toggle = IntVar()
        terms_con = Checkbutton(master, variable=toggle, onvalue=1, text='I agree to the ATM Machine Terms & Conditions.',font=(self.font,10), command=PopUp)
        terms_con.grid(row=8,column=1,padx=120,pady=320)
   
        def Register():
            reg_data = []
            
            reg_id = Id_entry.get()
            reg_name = Fullname_entry.get()
            reg_pw = Password_entry.get()
            reg_no = Num_entry.get()
            reg_age = Age_entry.get()
            reg_bal = Bal_entry.get()
            reg_gender = Gender.get()
            checked = toggle.get()

            xl = openpyxl.load_workbook("g9db.xlsx")
            data = xl.active
            acct_column = data['A']
            
            account = False
            for x in acct_column: 
                if x.value == reg_id:
                    account = True
            
            if not ((reg_id == '') or (reg_name == '') or (reg_pw== '') or (reg_age == '') or (reg_gender == '') or (checked == 0) or (reg_bal == '') or (reg_no == '')):
                if not(account):
                    if reg_pw.isdigit() and reg_bal.isdigit() and reg_no.isdigit():
                        if (len(reg_pw) == 5):
                            if (len(reg_no) == 11):
                                if int(reg_age) > 17:
                                    
                                    reg_data.append(Id_entry.get())
                                    reg_data.append(Fullname_entry.get())
                                    reg_data.append(Password_entry.get())
                                    reg_data.append(Num_entry.get())
                                    reg_data.append(Age_entry.get())
                                    reg_data.append(Bal_entry.get())
                                    reg_data.append(Gender.get())

                                    # open xl
                                    xl = openpyxl.load_workbook("g9db.xlsx")
                                    data = xl.active
                                    data.append(reg_data)
                                    # saveeee
                                    xl.save("g9db.xlsx")

                                    tkinter.messagebox.showinfo("Successful","Registration Completed!")
                                    
                                else:
                                    tkinter.messagebox.showinfo("Warning","You must 18+ to register.")
                            else:
                                tkinter.messagebox.showinfo("Warning","Contact Number must be 11 digits.")
                        else:
                            tkinter.messagebox.showinfo("Warning","For PASSWORD: Please enter 5 DIGITS only.")
                    else:
                        tkinter.messagebox.showinfo("Warning","For PASSWORD: Please enter 5 DIGITS only.")
                else:
                    tkinter.messagebox.showinfo("Failed","Account already exists.")
            else:
                tkinter.messagebox.showinfo("Failed","Please complete the registration")
            
        # back to log in page
        Button(master, text = "Log in", width = 5, font = ('Times New Roman', 15), bg = self.gray, fg = self.white, bd = 0, command = lambda: self.__init__(master)).place(x=150,y=350) 
        #for register button
        Button(master, text = "Register", width = 5, font = ('Times New Roman', 15), bg = self.green, fg = self.white, bd = 0, command=Register).place(x=280,y=350) 

        root.mainloop()

#################### option page, after login ####################
    def optionPg(self, master):
        # initializing frame 
        frame = Frame(root,width=600,height=500)
        frame.place(x=0,y=0)
        master.geometry('500x400')
        master.resizable(0,0)
        master.title("DSA Bank")

        # option page label
        label = Label(master, width = 43, text = "Choose a Transaction", font = (self.font, 25, 'bold'), background = '#d9d9d9', foreground = self.green)
        label.place(x = -175, y = 55)

        # note text, log out to cancel transaction
        label2 = Label(master, width = 43, text = "Choose LOG OUT if you\'d like to exit.", font = (self.font, 14), background = self.green, foreground = 'white', padx = 150, pady = 10)
        label2.place(x = -165, y = 100)

        # balance inquiry button
        bal_btn = Button(master, text = "Balance Inquiry", width = 23, height = 2, font = ('Times New Roman', 15), bg = self.green, fg = self.white,
        borderwidth = 0, command = lambda: self.balanceInq(master))
        bal_btn.place(x = -12, y = 175)

        # withdraw button
        withdraw_btn = Button(master, text = "Withdraw", width = 23, height = 2, font = ('Times New Roman', 15), bg = self.green, fg = self.white,
        borderwidth = 0, command = lambda: self.withdrawPg(master, self.user_infos[5]))
        withdraw_btn.place(x = -12, y = 245)

        # deposit button
        deposit_btn = Button(master, text = "Deposit", width = 23, height = 2, font = ('Times New Roman', 15), bg = self.green, fg = self.white,
        borderwidth = 0, command = lambda: self.depositPg(master, self.user_infos[5]))
        deposit_btn.place(x = 278, y = 175)

        # change password button
        changepass_btn = Button(master, text = "Change Password", width = 23, height = 2, font = ('Times New Roman', 15), bg = self.green, fg = self.white,
        borderwidth = 0, command = lambda: self.changePassPg(master))
        changepass_btn.place(x = 278, y = 245)

        # log out button, go back to login page weeeee
        logout_btn = Button(master, text = "Log out", width = 23, height = 2, font = ('Times New Roman', 15), bg = self.red, fg = self.white,
        borderwidth = 0, command = lambda: self.logout())
        logout_btn.place(x = 278, y = 315)

        root.mainloop()

#################### log out function ####################
    def logout(self):
        # reset program datas
        self.user_infos = []
        self.__init__(root)

#################### balance inquiry page ####################
    def balanceInq(self, master):
        # initializing frame 
        frame = Frame(root,width=600,height=500)
        frame.place(x=0,y=0)
        master.geometry('500x400')
        master.resizable(0,0)
        master.title("DSA Bank")

        label = Label(master, width = 43, text = "Balance Inquiry", pady = 5, font = (self.font, 16, 'bold'), background = self.red, foreground = "white")
        label.place(x = 10, y = 45)

        # account no label
        acctlbl = Label(master, text = "Account ID:", font = (self.font, 14))
        acctlbl.place(x = 150, y = 120)

        # show account no
        acctno = Label(master, text = self.user_infos[0], font = (self.font, 14, 'bold'))
        acctno.place(x = 275, y = 120)

        # balance label
        acctlbl = Label(master, text = "Balance:", font = (self.font, 14))
        acctlbl.place(x = 150, y = 170)

        # show acct balance
        acctno = Label(master, text = self.user_infos[5], font = (self.font, 14, 'bold'))
        acctno.place(x = 275, y = 170)

        # note after viewing account balance
        notelbl = Label(master, text = "After reviewing your account balance, \nselect CONTINUE.", font = (self.font, 14))
        notelbl.place(x = 100, y = 250)

        # continue button, continue to option page
        cont_btn = Button(master, text = "Continue", width = 10, font = ('Times New Roman', 15), bg = self.green, fg = self.white,
        borderwidth = 0, command = lambda: self.optionPg(master))
        cont_btn.place(x = 278, y = 315)

        root.mainloop()

#################### deposit page #################### 
    def depositPg(self, master, balance):
        # initializing frame
        frame = Frame(root,width=600,height=500)
        frame.place(x=0,y=0)
        master.geometry('500x400')
        master.resizable(0,0)
        
        master.title("DSA ATM MACHINE")
        frame.configure(bg="#d9d9d9")
        amount = StringVar()
        
        amount = StringVar()

        acca = 0
        acca = acca + int(balance)

        def deposit():
            try:
                amo = (amount.get())
                bal = acca + int(amo)

                # self.user_infos[7] = bal
                self.user_infos.pop(5)
                self.user_infos.insert(5, bal)

                # connect to db
                xl = openpyxl.load_workbook(self.xlfile)
                data = xl.active

                change = 'F'+str(row_no)
                data[change].value = bal

                xl.save(self.xlfile)

                tkinter.messagebox.showinfo("Successful", "Deposit successful!")

            except  ValueError:
                tkinter.messagebox.showerror("Error", "Please enter an amount only!")

        depacc_Label = Label (width = 43, text = "Deposit",pady = 5, font = ("Times New Roman", 16, 'bold'), background="#d63a3a", foreground="white")
        depacc_Label.grid(row = 1, column = 0, columnspan = 3, pady = 30, padx = 10)

        depentlbl = Label(text = "Enter an amount to deposit",font = ("Times New Roman", 15), bg ="#d9d9d9", anchor = 'w')
        depentlbl.place(x=155, y = 90)

        depEntry = Entry (font = ("Times New Roman", 14), textvariable =amount, fg = "black", bg = "white", bd=5, insertwidth=4, justify='right')
        depEntry.place(x=150, y = 120, width=250,height=50)

        # result of the amount deposited
        depResult = Label( font = ("Times New Roman", 14, 'bold'), fg='black', bg = "#d9d9d9")
        depResult.place(x = 160, y = 180)

        # deposit button
        depbutton = Button(master, text = "Deposit", width = 8, font = ('Times New Roman', 15), bg = self.green, fg = self.white, borderwidth=0, command = deposit)
        depbutton.place(x = 160, y = 220)

        # exit button
        exbutton = Button(master, text = "Cancel", width = 8, font = ('Times New Roman', 15), bg = self.red, fg = self.white, borderwidth = 0, command = lambda: self.optionPg(master))
        exbutton.place(x=280, y = 220)

        root.mainloop()

#################### withdraw page #################### 
    def withdrawPg(self, master, balance):
        # initializing frame
        frame = Frame(root,width=600,height=500)
        frame.place(x=0,y=0)
        master.geometry('500x400')
        master.resizable(0,0)
        
        master.title("DSA ATM MACHINE")
        frame.configure(bg="#d9d9d9")
        amount = StringVar()
        
        amount = StringVar()

        acca = 0
        acca = acca + int(balance)

        def withdraw():
            try:
                amo = (amount.get())

                if int(amo) > int(balance):
                    tkinter.messagebox.showerror("Failed", "You don't have that kind of balance. :(")
                else:
                    bal = acca - int(amo)
                    # self.user_infos[7] = bal
                    self.user_infos.pop(5)
                    self.user_infos.insert(5, int(bal))

                    # connect to db
                    xl = openpyxl.load_workbook(self.xlfile)
                    data = xl.active

                    change = 'F'+str(row_no)
                    data[change].value = bal

                    xl.save(self.xlfile)

                    tkinter.messagebox.showinfo("Successful", "Withdraw successful!")

            except  ValueError:
                tkinter.messagebox.showerror("Error", "Please enter an amount only!")

        withdraw_Label = Label (width = 43, text = "Withdraw",pady = 5, font = ("Times New Roman", 16, 'bold'), background="#d63a3a", foreground="white")
        withdraw_Label.grid(row = 1, column = 0, columnspan = 3, pady = 30, padx = 10)

        withdrawlbl = Label(text = "Enter an amount to withdraw",font = ("Times New Roman", 15), bg ="#d9d9d9", anchor = 'w')
        withdrawlbl.place(x=155, y = 90)

        withdrawEntry = Entry (font = ("Times New Roman", 14), textvariable =amount, fg = "black", bg = "white", bd=5, insertwidth=4, justify='right')
        withdrawEntry.place(x=150, y = 120, width=250,height=50)

        # result of the amount deposited
        withdrawResult = Label( font = ("Times New Roman", 14, 'bold'), fg='black', bg = "#d9d9d9")
        withdrawResult.place(x = 160, y = 180)

        # withdraw button
        withdrawbutton = Button(master, text = "Withdraw", width = 8, font = ('Times New Roman', 15), bg = self.green, fg = self.white, borderwidth=0, command = withdraw)
        withdrawbutton.place(x = 160, y = 220)

        # exit button
        exbutton = Button(master, text = "Cancel", width = 8, font = ('Times New Roman', 15), bg = self.gray, fg = self.white, borderwidth = 0, command = lambda: self.optionPg(master))
        exbutton.place(x=280, y = 220)

        root.mainloop()

#################### change password page #################### 
    def changePassPg(self, master):
        # initializing frame 
        frame = Frame(root,width=600,height=500)
        frame.place(x=0,y=0)
        master.geometry('500x400')
        master.resizable(0,0)
        master.title("DSA Bank")

        # change password page label
        label = Label(master, width = 43, text = "Change Password", pady = 5, font = (self.font, 16, 'bold'), background = self.red, foreground = "white")
        label.place(x = 10, y = 45)

        # account no label
        acctlbl = Label(master, text = "Current Password:", font = (self.font, 13))
        acctlbl.place(x = 80, y = 120)

        # show account no
        input_currentpass = Entry(width = 14, show = "*", font = ('Times New Roman', 20), bd = 0 )
        input_currentpass.place(x = 235,y = 115)

        # balance label
        acctlbl = Label(master, text = "New Password:", font = (self.font, 14))
        acctlbl.place(x = 80, y = 175)

        # show acct balance
        input_newpass = Entry(width = 14,show = "*",font = ('Times New Roman', 20), bd = 0 )
        input_newpass.place(x = 235,y = 170)

        # save password button
        savepass_btn = Button(master, text = "Save Password", width = 14, font = ('Times New Roman', 15), bg = self.green, fg = self.white,
        borderwidth = 0, command = lambda: checkPass(row_no))
        savepass_btn.place(x = 180, y = 220)

        def checkPass(row_no):
            # initializing frame 
            check_currentpass = input_currentpass.get()
            check_newpass = input_newpass.get()
            user_infos_pass = str(self.user_infos[2])

            # # check if not empty
            if (len(check_currentpass)) > 0 and (len(check_newpass) > 0):
                # check if digit
                if (len(check_currentpass)) == 5 and (len(check_newpass) == 5):
                    # check if digit, digits pass only ATM!!!
                    if check_currentpass.isdigit() and check_newpass.isdigit():
                        # check if user info & input are the same
                        if user_infos_pass == check_currentpass:
                            # check if new pass /= current pass
                            if user_infos_pass != check_newpass:
                                # open xl
                                xl = openpyxl.load_workbook("g9db.xlsx")
                                data = xl.active

                                # self.user_infos[8] = check_newpass
                                self.user_infos.pop(2)
                                self.user_infos.insert(2, check_newpass)

                                # changing password from I column in excel
                                change = 'C'+str(row_no)    # change = column value :)
                                data[change].value = check_newpass

                                # saveeee
                                xl.save("g9db.xlsx")

                                tkinter.messagebox.showinfo('Successful','Password Saved!')
                            else: 
                                tkinter.messagebox.showinfo('Failed','New password is the same as the current password.')
                        else:
                            tkinter.messagebox.showinfo('Failed','Current Password is incorrect.')
                    else:
                        tkinter.messagebox.showinfo('Failed','Enter number/s only.')
                else:
                    tkinter.messagebox.showinfo('Failed','Enter 5 digits for your pin.')

            # caution if entry box are empty        
            elif (len(check_currentpass)) == 0 or (len(check_newpass) == 0):
                tkinter.messagebox.showinfo('Failed','Current Password/New Password is empty.')   

        # note after chaging the password
        notelbl = Label(master, text = "Simply click CONTINUE once you've \nmodified your password (or not).", font = ("Times New Roman", 14))
        notelbl.place(x = 100, y = 260)

        # continue button to option page
        cont_btn = Button(master, text = "Continue", width = 8, font = ('Times New Roman', 15), bg = self.green, fg = self.white,
        borderwidth = 0, command = lambda: self.optionPg(master))
        cont_btn.place(x = 288, y = 315)

        root.mainloop()

ATMapp(root)
root.mainloop()