
from tkinter import *
import tkinter.messagebox
from tkinter import messagebox
import openpyxl
root = Tk()

class ATMapp:

    # import image
    bal_img = PhotoImage(file='button_balance-inquiry.png')
    withdraw_img = PhotoImage(file='button_withdraw.png')
    deposit_img = PhotoImage(file='button_deposit.png')
    changepass_img = PhotoImage(file='button_change-password.png')
    logout_img = PhotoImage(file='button_log-out.png')
    cont_img = PhotoImage(file='button_continue.png')
    savepass_img = PhotoImage(file = 'button_save-password.png')
    login_img = PhotoImage(file = 'button_log-in.png')
    reg_img = PhotoImage(file = 'button_register.png')
    reg2_img = PhotoImage(file = 'button_register (1).png')
    dep_img = PhotoImage(file ='button_depo.png')
    cancel_img = PhotoImage(file = 'button_cancel.png')

    xlfile = "g9db.xlsx"    # file name
    green = '#3aa15c' # color
    red = '#d63a3a'   # color
    font = "Times New Roman"    # font
    

    # where user's info will be appended, can be use in whole class! ;O
    user_infos = []

#################### login page, first page ####################
    def __init__(self, master):
        # initializing frame
        frame = Frame(root,width = 600,height = 500)
        frame.place(x=0,y=0)
        master.geometry('500x400')
        master.resizable(0,0)
        master.title("DSA Bank")
        
        # page login label
        label = Label(master, width = 43, text = "Log in", pady = 5, font = (self.font, 16, 'bold'), background = self.red, foreground = "white")
        label.place(x = 10, y = 45)

        # username label
        unlbl = Label(master, text = "Username:", font = (self.font, 13))
        unlbl.place(x = 120, y = 120)

        # entry box for username
        input_username = Entry(width = 14,font = ('Times New Roman', 20), bd = 0 )
        input_username.place(x = 235,y = 115)

        # password label
        pwlbl = Label(master, text = "Password:", font = (self.font, 14))
        pwlbl.place(x = 120, y = 175)

        # entry box for password
        input_userpass = Entry(width = 14,show = "*",font = ('Times New Roman', 20), bd = 0 )
        input_userpass.place(x = 235,y = 170)

        # log in button
        login_btn = Button(master, image = self.login_img,
        borderwidth = 0, command = lambda: self.importdata(input_username, input_userpass))
        login_btn.place(x = 350, y = 220)

        # register button
        register_btn = Button(master, image = self.reg_img,
        borderwidth = 0, command = lambda: self.registerPg(master))
        register_btn.place(x = 300, y = 265)


    def getdata(self, user, row):
        # user's specific row in database
        global row_no
        row_no = row
            # user from database append to user_infos to use in whole program
        for info in user:
            self.user_infos.append(info)

        # successful log in! go to option page :)
        self.optionPg(root)

    def importdata(self, get_username, get_password):   
        # counter for row
        counter = 0

        # where row of data will be appended
        db_user = [] 

        # from login page user inputs: username & password
        # using .get() function to turn into value/s
        input_un = get_username.get()
        input_pw = get_password.get()

        # opening a excel py
        xl = openpyxl.load_workbook(self.xlfile)
        data = xl.active
        rws = data.iter_rows(min_row = 1, max_col= 9, max_row = None, min_col = 1, values_only = True)
        
        # iterating rows in excel
        for i, j, k, l, m, n, o, p, q in rws:
            counter = counter + 1
            if o == input_un:
                # if user is true, append data to db_user list
                db_user.extend([i, j, k, l, m, n, o, p, q])
                break

        # checking if login input pw == database pw
        if (len(db_user)) == 0 or (len(input_pw)) == 0:
            tkinter.messagebox.showinfo('Failed','Invalid Login Username or Password')
        elif (len(db_user[8])) > 0:
            userpass = str(db_user[8])
            if input_pw == userpass:
                # if == call getdata() to save THAT row/user info's
                self.getdata(db_user, counter)
            else:
                tkinter.messagebox.showinfo('Failed','Invalid Login Username or Password')
        else:
            tkinter.messagebox.showinfo('Failed','Invalid Login Username or Password')
        row_no = counter
        root.mainloop()

#################### registration page ####################
    def registerPg(self, master):
        # initializing frame
        frame = Frame(root,width=600,height=500)
        frame.place(x=0,y=0)
        master.geometry('500x400')
        master.resizable(0,0)

        myLabel = Label(self, text = "Registration Form", width=20, font=("bold", 20))
        myLabel.place(x=97,y=20)


        #Id number
        LblId = Label(self, text="ID Number",width=20,font=("bold", 10))  
        LblId.place(x=80,y=100) 

        Id_entry = IntVar        
        Id_entry = Entry(self)  
        Id_entry.place(x=250,y=100,width=150)   
   
        
        #fullname        
        LblFullname = Label(self, text="Full Name",width=20,font=("bold", 10))  
        LblFullname.place(x=80,y=150) 

        Fullname_entry = StringVar        
        Fullname_entry = Entry(self)  
        Fullname_entry.place(x=250,y=150,width=150)  

        
        #password        
        LblPassword = Label(self, text="Password",width=20,font=("bold", 10))  
        LblPassword.place(x=80,y=200) 

        Password_entry = StringVar        
        Password_entry = Entry(self)  
        Password_entry.place(x=250,y=200,width=150)  

        #contact number
        LblNum = Label(self, text="Contact number",width=20,font=("bold", 10))  
        LblNum.place(x=95,y=250) 

        Num_entry = IntVar         
        Num_entry = Entry(self)  
        Num_entry.place(x=250,y=250,width=150) 

        
        #email        
        LblEmail = Label(self, text="Email",width=20,font=("bold", 10))  
        LblEmail.place(x=66,y=300) 

        Email_entry = StringVar         
        Email_entry = Entry(self)  
        Email_entry.place(x=250,y=300,width=150) 


        #balance        
        LblBal = Label(self, text="Balance",width=20,font=("bold", 10))  
        LblBal.place(x=71,y=350) 

        Bal_entry = DoubleVar         
        Bal_entry = Entry(self)  
        Bal_entry.place(x=250,y=350,width=150) 

        
        #gender        
        LblGender = Label(self, text="Gender",width=20,font=("bold", 10))  
        LblGender.place(x=69,y=400)
        
        Gender = IntVar()
        Radiobutton(self, text="Male",padx = 10, variable=Gender, value=1).place(x=243,y=400)
        Radiobutton(self, text="Female",padx = 20, variable=Gender, value=2).place(x=300,y=400)  


        #terms and condition
        terms_con = Checkbutton(self, text='I agree to the ATM Machine Terms & Conditions.',font=("Bold",10))
        terms_con.grid(row=8,column=1,padx=120,pady=450)
            
        def Register():
            reg_data = []
            reg_data.append(Fullname_entry.get())
            reg_data.append(Password_entry.get())
            reg_data.append(Num_entry.get())
            reg_data.append(Email_entry.get())
            reg_data.append(Gender.get())
            print(reg_data)
            tkinter.messagebox.showinfo("Registration","Successfully Registered")
                
        #for register button
        Button(master, text='Register',width=20,bg='black',fg='white', command=Register).place(x=180,y=360) 

        root.mainloop()

#################### option page, after login ####################
    def optionPg(self, master):
        # initializing frame 
        frame = Frame(root,width=600,height=500)
        frame.place(x=0,y=0)
        master.geometry('500x400')
        master.resizable(0,0)
        master.title("DSA Bank")

        # option page label
        label = Label(master, width = 43, text = "Choose a Transaction", font = (self.font, 25, 'bold'), background = '#d9d9d9', foreground = self.green)
        label.place(x = -175, y = 55)

        # note text, log out to cancel transaction
        label2 = Label(master, width = 43, text = "Choose LOG OUT if you\'d like to exit.", font = (self.font, 14), background = self.green, foreground = 'white', padx = 150, pady = 10)
        label2.place(x = -165, y = 100)

        # balance inquiry button
        bal_btn = Button(master, image = self.bal_img,
        borderwidth = 0, command = lambda: self.balanceInq(master))
        bal_btn.place(x = -12, y = 175)

        # withdraw button
        withdraw_btn = Button(master, image = self.withdraw_img,
        borderwidth = 0)
        withdraw_btn.place(x = -12, y = 245)

        # deposit button
        deposit_btn = Button(master, image = self.deposit_img,
        borderwidth = 0, command = lambda: self.depositPg(master, self.user_infos[7]))
        deposit_btn.place(x = 278, y = 175)

        # change password button
        changepass_btn = Button(master, image = self.changepass_img,
        borderwidth = 0, command = lambda: self.changePassPg(master))
        changepass_btn.place(x = 278, y = 245)

        # log out button, go back to login page weeeee
        logout_btn = Button(master, image = self.logout_img,
        borderwidth = 0, command = lambda: self.logout())
        logout_btn.place(x = 278, y = 315)

        root.mainloop()

#################### log out function ####################
    def logout(self):
        # reset program datas
        self.user_infos = []
        self.__init__(root)

#################### balance inquiry page ####################
    def balanceInq(self, master):
        # initializing frame 
        frame = Frame(root,width=600,height=500)
        frame.place(x=0,y=0)
        master.geometry('500x400')
        master.resizable(0,0)
        master.title("DSA Bank")


        label = Label(master, width = 43, text = "Balance Inquiry", pady = 5, font = (self.font, 16, 'bold'), background = self.red, foreground = "white")
        label.place(x = 10, y = 45)

        # account no label
        acctlbl = Label(master, text = "Account No.:", font = (self.font, 14))
        acctlbl.place(x = 150, y = 120)

        # show account no
        acctno = Label(master, text = self.user_infos[0], font = (self.font, 14, 'bold'))
        acctno.place(x = 275, y = 120)

        # balance label
        acctlbl = Label(master, text = "Balance:", font = (self.font, 14))
        acctlbl.place(x = 150, y = 170)

        # show acct balance
        acctno = Label(master, text = self.user_infos[7], font = (self.font, 14, 'bold'))
        acctno.place(x = 275, y = 170)

        # note after viewing account balance
        notelbl = Label(master, text = "After reviewing your account balance, \nselect CONTINUE.", font = (self.font, 14))
        notelbl.place(x = 100, y = 250)

        # continue button, continue to option page
        cont_btn = Button(master, image = self.cont_img,
        borderwidth = 0, command = lambda: self.optionPg(master))
        cont_btn.place(x = 278, y = 315)

        root.mainloop()

#################### deposit page #################### 
    def depositPg(self, master, balance):
        # initializing frame
        frame = Frame(root,width=600,height=500)
        frame.place(x=0,y=0)
        master.geometry('500x400')
        master.resizable(0,0)
        
        master.title("DSA ATM MACHINE")
        frame.configure(bg="#d9d9d9")
        amount = StringVar()
        
        amount = StringVar()

        acca = 0
        acca = acca + balance

        def deposit():
            try:
                amo = (amount.get())
                bal = acca + int(amo)
                depEntry.config(text=(f"Current Balance: {bal}"))

                # self.user_infos[7] = bal
                self.user_infos.pop(7)
                self.user_infos.insert(7, bal)

                # connect to db
                xl = openpyxl.load_workbook(self.xlfile)
                data = xl.active

                change = 'H'+str(row_no)
                data[change].value = bal

                xl.save(self.xlfile)
            except  ValueError:
                messagebox.showerror("Error", "Please enter an amount only!")



        depacc_Label = Label (width = 43, text = "Deposit Account",pady = 5, font = ("Times New Roman", 16, 'bold'), background="#d63a3a", foreground="white")
        depacc_Label.grid(row = 1, column = 0, columnspan = 3, pady = 30, padx = 10)

        depentlbl = Label(text = "Enter an amount to deposit",font = ("Times New Roman", 15), bg ="#d9d9d9", anchor = 'w')
        depentlbl.place(x=155, y = 90)

        depEntry = Entry (font = ("Times New Roman", 14), textvariable =amount, fg = "black", bg = "white", bd=5, insertwidth=4, justify='right')
        depEntry.place(x=150, y = 120, width=250,height=50)

        # result of the amount deposited
        depResult = Label( font = ("Times New Roman", 14, 'bold'), fg='black', bg = "#d9d9d9")
        depResult.place(x = 160, y = 180)

        # deposit button
        depbutton = Button(master, image = self.dep_img, borderwidth=0, command = deposit)
        depbutton.place(x = 150, y = 220)

        # exit button
        exbutton = Button(master, image = self.cancel_img, borderwidth = 0, command = lambda: self.optionPg(master))
        exbutton.place(x=270, y = 220)


        root.mainloop()

#################### change password page #################### 
    def changePassPg(self, master):
        # initializing frame 
        frame = Frame(root,width=600,height=500)
        frame.place(x=0,y=0)
        master.geometry('500x400')
        master.resizable(0,0)
        master.title("DSA Bank")

        # change password page label
        label = Label(master, width = 43, text = "Change Password", pady = 5, font = (self.font, 16, 'bold'), background = self.red, foreground = "white")
        label.place(x = 10, y = 45)

        # account no label
        acctlbl = Label(master, text = "Current Password:", font = (self.font, 13))
        acctlbl.place(x = 80, y = 120)

        # show account no
        input_currentpass = Entry(width = 14, show = "*", font = ('Times New Roman', 20), bd = 0 )
        input_currentpass.place(x = 235,y = 115)

        # balance label
        acctlbl = Label(master, text = "New Password:", font = (self.font, 14))
        acctlbl.place(x = 80, y = 175)

        # show acct balance
        input_newpass = Entry(width = 14,show = "*",font = ('Times New Roman', 20), bd = 0 )
        input_newpass.place(x = 235,y = 170)

        # save password button
        savepass_btn = Button(master, image = self.savepass_img,
        borderwidth = 0, command = lambda: checkPass(row_no))
        savepass_btn.place(x = 180, y = 220)

        def checkPass(row_no):
            # initializing frame 
            check_currentpass = input_currentpass.get()
            check_newpass = input_newpass.get()
            user_infos_pass = str(self.user_infos[8])

            # # check if not empty
            if (len(check_currentpass)) > 0 and (len(check_newpass) > 0):
                # check if digit
                if (len(check_currentpass)) == 5 and (len(check_newpass) == 5):
                    # check if digit, digits pass only ATM!!!
                    if check_currentpass.isdigit() and check_newpass.isdigit():
                        # check if user info & input are the same
                        if user_infos_pass == check_currentpass:
                            # check if new pass /= current pass
                            if user_infos_pass != check_newpass:
                                # open xl
                                xl = openpyxl.load_workbook("g9db.xlsx")
                                data = xl.active

                                # self.user_infos[8] = check_newpass
                                self.user_infos.pop(8)
                                self.user_infos.insert(8, check_newpass)

                                # changing password from I column in excel
                                change = 'I'+str(row_no)    # change = column value :)
                                data[change].value = check_newpass

                                # saveeee
                                xl.save("g9db.xlsx")

                                tkinter.messagebox.showinfo('Successful','Password Saved!')
                            else: 
                                tkinter.messagebox.showinfo('Failed','New password is the same as the current password.')
                        else:
                            # print((str(user_infos[8])).isdigit(), check_currentpass.isdigit())
                            tkinter.messagebox.showinfo('Failed','Current Password is incorrect.')
                    else:
                        tkinter.messagebox.showinfo('Failed','Enter number/s only.')
                else:
                    tkinter.messagebox.showinfo('Failed','Enter 5 digits for your pin.')

            # caution if entry box are empty        
            elif (len(check_currentpass)) == 0 or (len(check_newpass) == 0):
                tkinter.messagebox.showinfo('Failed','Current Password/New Password is empty.')   

        # note after chaging the password
        notelbl = Label(master, text = "After changing your password, select CONTINUE.", font = ("Times New Roman", 14))
        notelbl.place(x = 80, y = 280)

        # continue button to option page
        cont_btn = Button(master, image = self.cont_img,
        borderwidth = 0, command = lambda: self.optionPg(master))
        cont_btn.place(x = 278, y = 315)

        root.mainloop()

ATMapp(root)
root.mainloop()
