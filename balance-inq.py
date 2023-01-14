import os
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

# trial uli for acct no and balance
def getData():
    acct_no = 987456321
    bal_inq = 100
    count = 0

    # creating file for trial hoho
    # wb = Workbook()
    # wb.save(filename = 'nag-iisa-lang-to.xlsx')

    xl = openpyxl.load_workbook("nag-iisa-lang-to.xlsx")
  
    data = xl.active

    acctno_column = data['A']           # get data from A column


    # parang may ibang way pa na mas maganda to find data? ayaw lang i-type ng daliri ko 
    for x in range(len(acctno_column)):
        if acct_no == acctno_column[x].value:
            print("hello")
        count = count + 1 

    for row in data.iter_rows(min_row = count, max_col= None, max_row = count, values_only = True):
        for infos in row:
            if infos == bal_inq:
                print("korik dzai")


    balanceInq(acct_no, bal_inq)

def balanceInq(acct_no, bal_inq):
    x = acct_no
    m = bal_inq
    print("Balance Inquiry\n")

    print("\nAccount Number: ", x)

    print("Your balance: ", m)

    print("\n\tRecord saved!")
    print("After reviewing your account balance,\nenter [C] to continue.")
    
getData()